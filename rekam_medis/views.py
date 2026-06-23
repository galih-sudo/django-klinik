from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import models
from django.http import HttpResponse
from django.template.loader import get_template
from .models import Pasien, Obat, Kunjungan, RekamMedis, ICD10, Profile
from .forms import PasienForm, RekamMedisForm
from .decorators import dokter_required, staf_required
from django.core.paginator import Paginator

# PDF
from weasyprint import HTML

# Excel
import openpyxl
from openpyxl import Workbook
import pandas as pd
from io import BytesIO

# ========== DASHBOARD ==========
@login_required
def dashboard(request):
    today = timezone.now().date()
    total_pasien = Pasien.objects.filter(aktif=True).count()
    total_obat = Obat.objects.count()
    total_kunjungan = Kunjungan.objects.filter(tanggal=today).count()
    pasien_terbaru = Pasien.objects.filter(aktif=True).order_by('-id')[:10]

    context = {
        'today': today.strftime('%Y-%m-%d'),
        'total_pasien': total_pasien,
        'total_obat': total_obat,
        'total_kunjungan': total_kunjungan,
        'pasien_terbaru': pasien_terbaru,
    }
    return render(request, 'rekam_medis/dashboard.html', context)

# ========== PASIEN (Staf & Dokter bisa) ==========
@login_required
@staf_required
def daftar_pasien(request):
    pasien_list = Pasien.objects.filter(aktif=True).order_by('-id')

    paginator = Paginator(pasien_list, 10)  # 10 pasien per halaman
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }
    return render(request, 'rekam_medis/daftar_pasien.html', context)

@login_required
@staf_required
def tambah_pasien(request):
    if request.method == 'POST':
        form = PasienForm(request.POST)
        if form.is_valid():
            pasien = form.save(commit=False)
            tahun = timezone.now().strftime('%y')
            count = Pasien.objects.count() + 1
            pasien.no_rm = f"{tahun}-{count:04d}"
            pasien.save()
            messages.success(request, f'✅ Pasien {pasien.nama} berhasil ditambahkan! No RM: {pasien.no_rm}')
            return redirect('dashboard')
    else:
        form = PasienForm()
    return render(request, 'rekam_medis/tambah_pasien.html', {'form': form})

@login_required
@staf_required
def cari_pasien(request):
    keyword = request.GET.get('q', '').strip()
    jenis_kelamin = request.GET.get('jenis_kelamin', '')
    tgl_awal = request.GET.get('tgl_awal', '')
    tgl_akhir = request.GET.get('tgl_akhir', '')
    status_aktif = request.GET.get('status_aktif', '')

    pasien_list = Pasien.objects.all()

    if keyword:
        pasien_list = pasien_list.filter(
            models.Q(nama__icontains=keyword) |
            models.Q(no_rm__icontains=keyword) |
            models.Q(alamat__icontains=keyword)
        )

    if jenis_kelamin:
        pasien_list = pasien_list.filter(jenis_kelamin=jenis_kelamin)

    if tgl_awal:
        pasien_list = pasien_list.filter(tgl_lahir__gte=tgl_awal)
    if tgl_akhir:
        pasien_list = pasien_list.filter(tgl_lahir__lte=tgl_akhir)

    if status_aktif == 'aktif':
        pasien_list = pasien_list.filter(aktif=True)
    elif status_aktif == 'tidak_aktif':
        pasien_list = pasien_list.filter(aktif=False)

    context = {
        'pasien_list': pasien_list,
        'keyword': keyword,
        'jenis_kelamin': jenis_kelamin,
        'tgl_awal': tgl_awal,
        'tgl_akhir': tgl_akhir,
        'status_aktif': status_aktif,
        'total': pasien_list.count(),
    }
    return render(request, 'rekam_medis/cari_pasien.html', context)

# ========== DOKTER ONLY ==========
@login_required
@dokter_required
def edit_pasien(request, pasien_id):
    pasien = get_object_or_404(Pasien, id=pasien_id)

    if request.method == 'POST':
        form = PasienForm(request.POST, instance=pasien)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Data pasien {pasien.nama} berhasil diupdate!')
            return redirect('daftar_pasien')
    else:
        form = PasienForm(instance=pasien)

    return render(request, 'rekam_medis/edit_pasien.html', {'form': form, 'pasien': pasien})

@login_required
@dokter_required
def rekam_medis(request, pasien_id):
    pasien = get_object_or_404(Pasien, id=pasien_id)
    icd_list = ICD10.objects.all().order_by('kode')

    if request.method == 'POST':
        form = RekamMedisForm(request.POST)
        if form.is_valid():
            rekam = form.save(commit=False)
            rekam.pasien = pasien
            rekam.save()

            Kunjungan.objects.create(
                rekam_medis=rekam,
                tanggal=timezone.now().date()
            )
            messages.success(request, f'✅ Rekam medis untuk {pasien.nama} berhasil disimpan!')
            return redirect('daftar_pasien')
    else:
        form = RekamMedisForm()

    context = {
        'form': form,
        'pasien': pasien,
        'icd_list': icd_list,
    }
    return render(request, 'rekam_medis/rekam_medis.html', context)

@login_required
@dokter_required
def buat_resep(request, rekam_medis_id):
    rekam = get_object_or_404(RekamMedis, id=rekam_medis_id)
    obat_list = Obat.objects.all().order_by('nama')

    if request.method == 'POST':
        obat_id = request.POST.get('obat_id')
        jumlah = int(request.POST.get('jumlah', 0))
        aturan = request.POST.get('aturan', '')

        obat = get_object_or_404(Obat, id=obat_id)

        if obat.stok >= jumlah:
            obat.stok -= jumlah
            obat.save()

            from .models import Resep
            Resep.objects.create(
                rekam_medis=rekam,
                obat=obat,
                jumlah=jumlah,
                aturan=aturan
            )
            messages.success(request, '✅ Resep berhasil dibuat!')
            return redirect('daftar_pasien')
        else:
            messages.error(request, f'❌ Stok {obat.nama} tidak cukup!')
            return render(request, 'rekam_medis/error_stok.html', {'obat': obat})

    context = {
        'rekam': rekam,
        'obat_list': obat_list,
    }
    return render(request, 'rekam_medis/buat_resep.html', context)

# ========== OBAT ==========
@login_required
@dokter_required
def kelola_obat(request):
    obat_list = Obat.objects.all().order_by('nama')
    return render(request, 'rekam_medis/kelola_obat.html', {'obat_list': obat_list})

@login_required
@dokter_required
def tambah_obat(request):
    if request.method == 'POST':
        nama = request.POST.get('nama')
        stok = int(request.POST.get('stok', 0))
        satuan = request.POST.get('satuan', 'tablet')

        Obat.objects.create(nama=nama, stok=stok, satuan=satuan)
        messages.success(request, f'✅ Obat {nama} berhasil ditambahkan!')
        return redirect('kelola_obat')

    return render(request, 'rekam_medis/tambah_obat.html')

@login_required
@dokter_required
def edit_obat(request, obat_id):
    obat = get_object_or_404(Obat, id=obat_id)

    if request.method == 'POST':
        obat.nama = request.POST.get('nama', obat.nama)
        obat.stok = int(request.POST.get('stok', obat.stok))
        obat.satuan = request.POST.get('satuan', obat.satuan)
        obat.save()
        messages.success(request, f'✅ Obat {obat.nama} berhasil diupdate!')
        return redirect('kelola_obat')

    return render(request, 'rekam_medis/edit_obat.html', {'obat': obat})

@login_required
@dokter_required
def hapus_obat(request, obat_id):
    obat = get_object_or_404(Obat, id=obat_id)
    nama = obat.nama
    obat.delete()
    messages.success(request, f'✅ Obat {nama} berhasil dihapus!')
    return redirect('kelola_obat')

# ========== ICD-10 ==========
@login_required
@dokter_required
def kelola_icd10(request):
    icd_list = ICD10.objects.all().order_by('kode')

    if request.method == 'POST':
        kode = request.POST.get('kode')
        nama = request.POST.get('nama_penyakit')
        kategori = request.POST.get('kategori', '')

        if kode and nama:
            ICD10.objects.create(
                kode=kode,
                nama_penyakit=nama,
                kategori=kategori
            )
            messages.success(request, f'✅ ICD-10 {kode} berhasil ditambahkan!')
            return redirect('kelola_icd10')

        hapus_id = request.POST.get('hapus_id')
        if hapus_id:
            icd = get_object_or_404(ICD10, kode=hapus_id)
            icd.delete()
            messages.success(request, f'✅ ICD-10 {hapus_id} berhasil dihapus!')
            return redirect('kelola_icd10')

    context = {
        'icd_list': icd_list,
    }
    return render(request, 'rekam_medis/kelola_icd10.html', context)

@login_required
@dokter_required
def tambah_icd10(request):
    if request.method == 'POST':
        kode = request.POST.get('kode')
        nama_penyakit = request.POST.get('nama_penyakit')
        kategori = request.POST.get('kategori', '')

        if kode and nama_penyakit:
            if ICD10.objects.filter(kode=kode).exists():
                messages.error(request, f'❌ Kode ICD-10 {kode} sudah ada!')
            else:
                ICD10.objects.create(
                    kode=kode,
                    nama_penyakit=nama_penyakit,
                    kategori=kategori
                )
                messages.success(request, f'✅ ICD-10 {kode} berhasil ditambahkan!')
            return redirect('kelola_icd10')

    return render(request, 'rekam_medis/tambah_icd10.html')

@login_required
@dokter_required
def edit_icd10(request, kode):
    icd = get_object_or_404(ICD10, kode=kode)

    if request.method == 'POST':
        icd.nama_penyakit = request.POST.get('nama_penyakit', icd.nama_penyakit)
        icd.kategori = request.POST.get('kategori', icd.kategori)
        icd.save()
        messages.success(request, f'✅ ICD-10 {kode} berhasil diupdate!')
        return redirect('kelola_icd10')

    context = {
        'icd': icd,
    }
    return render(request, 'rekam_medis/edit_icd10.html', context)

@login_required
@dokter_required
def hapus_icd10(request, kode):
    icd = get_object_or_404(ICD10, kode=kode)
    icd.delete()
    messages.success(request, f'✅ ICD-10 {kode} berhasil dihapus!')
    return redirect('kelola_icd10')

# ========== MANAJEMEN USER ==========
@login_required
@dokter_required
def kelola_user(request):
    users = User.objects.all().select_related('profile')
    return render(request, 'rekam_medis/kelola_user.html', {'users': users})

@login_required
@dokter_required
def tambah_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email', '')
        role = request.POST.get('role')

        if User.objects.filter(username=username).exists():
            messages.error(request, '❌ Username sudah ada!')
            return redirect('tambah_user')

        user = User.objects.create_user(username=username, password=password, email=email)
        Profile.objects.create(user=user, role=role)

        messages.success(request, f'✅ User {username} berhasil dibuat!')
        return redirect('kelola_user')

    return render(request, 'rekam_medis/tambah_user.html')

@login_required
@dokter_required
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile = user.profile

    if request.method == 'POST':
        user.email = request.POST.get('email', user.email)
        profile.role = request.POST.get('role', profile.role)
        profile.no_izin = request.POST.get('no_izin', profile.no_izin)
        profile.telepon = request.POST.get('telepon', profile.telepon)
        profile.alamat = request.POST.get('alamat', profile.alamat)

        user.save()
        profile.save()

        messages.success(request, f'✅ User {user.username} berhasil diupdate!')
        return redirect('kelola_user')

    return render(request, 'rekam_medis/edit_user.html', {'user': user, 'profile': profile})

@login_required
@dokter_required
def hapus_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, '❌ Tidak bisa menghapus akun sendiri!')
        return redirect('kelola_user')

    user.delete()
    messages.success(request, '✅ User berhasil dihapus!')
    return redirect('kelola_user')

# ========== CETAK SURAT ==========
@login_required
@dokter_required
def surat_sehat(request, pasien_id):
    pasien = get_object_or_404(Pasien, id=pasien_id)
    today = timezone.now().date()

    context = {
        'pasien': pasien,
        'today': today,
        'dokter': request.user,
    }

    template = get_template('rekam_medis/surat_sehat.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=surat_sehat_{pasien.no_rm}.pdf'

    HTML(string=html).write_pdf(response)
    return response

@login_required
@dokter_required
def surat_istirahat(request, pasien_id):
    pasien = get_object_or_404(Pasien, id=pasien_id)
    today = timezone.now().date()
    lama = request.GET.get('lama', '3')

    context = {
        'pasien': pasien,
        'today': today,
        'lama': lama,
        'dokter': request.user,
    }

    template = get_template('rekam_medis/surat_istirahat.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=surat_istirahat_{pasien.no_rm}.pdf'

    HTML(string=html).write_pdf(response)
    return response

# ========== EXPORT EXCEL ==========
@login_required
@dokter_required
def export_pasien_excel(request):
    pasien_list = Pasien.objects.all().order_by('-id')

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pasien"

    headers = ['No RM', 'Nama', 'Jenis Kelamin', 'Tgl Lahir', 'Alamat', 'No Telp', 'Pekerjaan', 'Alergi', 'Status']
    ws.append(headers)

    for p in pasien_list:
        ws.append([
            p.no_rm,
            p.nama,
            p.get_jenis_kelamin_display(),
            p.tgl_lahir.strftime('%Y-%m-%d') if p.tgl_lahir else '',
            p.alamat,
            p.no_telp,
            p.pekerjaan,
            p.alergi,
            'Aktif' if p.aktif else 'Tidak Aktif'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data_pasien.xlsx'
    wb.save(response)
    return response

@login_required
@dokter_required
def export_obat_excel(request):
    obat_list = Obat.objects.all().order_by('nama')

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Obat"

    headers = ['Nama Obat', 'Stok', 'Satuan']
    ws.append(headers)

    for o in obat_list:
        ws.append([o.nama, o.stok, o.satuan])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data_obat.xlsx'
    wb.save(response)
    return response

@login_required
@dokter_required
def export_icd10_excel(request):
    icd_list = ICD10.objects.all().order_by('kode')

    wb = Workbook()
    ws = wb.active
    ws.title = "Data ICD-10"

    headers = ['Kode', 'Nama Penyakit', 'Kategori']
    ws.append(headers)

    for i in icd_list:
        ws.append([i.kode, i.nama_penyakit, i.kategori])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data_icd10.xlsx'
    wb.save(response)
    return response

# ========== IMPORT EXCEL ==========
@login_required
@dokter_required
def import_pasien_excel(request):
    if request.method == 'POST' and request.FILES.get('file_excel'):
        file = request.FILES['file_excel']

        try:
            df = pd.read_excel(file)

            for index, row in df.iterrows():
                no_rm = str(row['No RM']) if pd.notna(row['No RM']) else None

                if no_rm and Pasien.objects.filter(no_rm=no_rm).exists():
                    continue

                pasien = Pasien(
                    no_rm=no_rm or f"{timezone.now().strftime('%y')}-{Pasien.objects.count() + 1:04d}",
                    nama=row['Nama'] if pd.notna(row['Nama']) else '',
                    jenis_kelamin=row['Jenis Kelamin'][0] if pd.notna(row['Jenis Kelamin']) and row['Jenis Kelamin'] else '',
                    tgl_lahir=row['Tgl Lahir'] if pd.notna(row['Tgl Lahir']) else None,
                    alamat=row['Alamat'] if pd.notna(row['Alamat']) else '',
                    no_telp=row['No Telp'] if pd.notna(row['No Telp']) else '',
                    pekerjaan=row['Pekerjaan'] if pd.notna(row['Pekerjaan']) else '',
                    alergi=row['Alergi'] if pd.notna(row['Alergi']) else '',
                    aktif=row['Status'] == 'Aktif' if pd.notna(row['Status']) else True
                )
                pasien.save()

            messages.success(request, '✅ Import data pasien berhasil!')

        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')

        return redirect('daftar_pasien')

    return render(request, 'rekam_medis/import_pasien.html')

# ========== LAPORAN ==========
@login_required
@dokter_required
def laporan_penyakit_terbanyak(request):
    from django.db.models import Count

    penyakit = ICD10.objects.annotate(
        total=Count('rekammedis')
    ).filter(total__gt=0).order_by('-total')[:10]

    context = {
        'penyakit': penyakit,
    }
    return render(request, 'rekam_medis/laporan_penyakit.html', context)

@login_required
@dokter_required
def laporan_obat_terbanyak(request):
    from django.db.models import Sum

    obat = Obat.objects.annotate(
        total_dipakai=Sum('resep__jumlah')
    ).filter(total_dipakai__gt=0).order_by('-total_dipakai')[:10]

    context = {
        'obat': obat,
    }
    return render(request, 'rekam_medis/laporan_obat.html', context)

@login_required
@dokter_required
def laporan_kunjungan(request):
    tgl_awal = request.GET.get('tgl_awal', '')
    tgl_akhir = request.GET.get('tgl_akhir', '')

    kunjungan = Kunjungan.objects.all().order_by('-tanggal')

    if tgl_awal:
        kunjungan = kunjungan.filter(tanggal__gte=tgl_awal)
    if tgl_akhir:
        kunjungan = kunjungan.filter(tanggal__lte=tgl_akhir)

    context = {
        'kunjungan': kunjungan,
        'tgl_awal': tgl_awal,
        'tgl_akhir': tgl_akhir,
        'total': kunjungan.count(),
    }
    return render(request, 'rekam_medis/laporan_kunjungan.html', context)